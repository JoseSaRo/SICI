from io import BytesIO
from datetime import datetime, timedelta
import base64
import csv
import re
from io import TextIOWrapper
from urllib.parse import quote, urlencode

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_POST
import qrcode
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Equipo, EquipoEliminado, MarcaEquipo, Movimiento, ResguardoEquipo, Responsable, SolicitudMarcaEquipo, SolicitudTipoEquipo, TipoEquipo, Ubicacion, UbicacionFisica


# Configuracion compartida por los flujos de inventario e importacion.
User = get_user_model()

RAM_OPTIONS = {
    "2 GB",
    "4 GB",
    "6 GB",
    "8 GB",
    "12 GB",
    "16 GB",
    "24 GB",
    "32 GB",
    "64 GB",
    "128 GB",
}

STORAGE_OPTIONS = {
    "120 GB SSD",
    "128 GB SSD",
    "240 GB SSD",
    "250 GB SSD",
    "256 GB SSD",
    "480 GB SSD",
    "500 GB HDD",
    "500 GB SSD",
    "512 GB SSD",
    "1 TB HDD",
    "1 TB SSD",
    "2 TB HDD",
    "2 TB SSD",
    "4 TB HDD",
}

EQUIPMENT_IMPORT_HEADERS = [
    "tipo_equipo",
    "marca",
    "modelo",
    "numero_serie",
    "procesador",
    "memoria_ram",
    "almacenamiento",
    "sistema_operativo",
    "mac_ethernet",
    "mac_wifi",
    "fecha_compra",
    "garantia_hasta",
    "ffoo_o_mesa",
    "ubicacion_fisica",
    "estado",
    "asignado_a",
    "observaciones",
]


# Preparacion de datos que consume la interfaz principal.
def format_date(value):
    return value.strftime("%d/%m/%Y") if value else "Sin registro"


def equipment_payload(equipo):
    return {
        "type": equipo.tipo,
        "typeId": equipo.tipo_equipo_id,
        "name": f"{equipo.marca} {equipo.modelo}".strip() or equipo.tipo,
        "brand": equipo.marca or "Sin marca",
        "serial": equipo.numero_serie,
        "createdAt": timezone.localtime(equipo.creado).isoformat(),
        "user": equipo.asignado_a or "Sin asignar",
        "front": equipo.ubicacion.nombre,
        "locationId": equipo.ubicacion_id,
        "physicalLocation": equipo.ubicacion_fisica.nombre if equipo.ubicacion_fisica else "Santa Luc\u00eda",
        "physicalLocationId": equipo.ubicacion_fisica_id,
        "status": equipo.get_estado_display(),
        "warranty": format_date(equipo.garantia_hasta),
        "cpu": equipo.procesador or "No aplica",
        "ram": equipo.memoria_ram or "No aplica",
        "storage": equipo.almacenamiento or "No aplica",
        "os": equipo.sistema_operativo or "No aplica",
        "macEthernet": equipo.mac_ethernet or "No aplica",
        "macWifi": equipo.mac_wifi or "No aplica",
        "invoiceUrl": equipo.factura.url if equipo.factura else "",
        "notes": equipo.observaciones or "Sin observaciones.",
        "history": [movement_payload(movimiento) for movimiento in equipo.movimientos.all()],
        "safekeeping": [safekeeping_payload(resguardo) for resguardo in equipo.resguardos.filter(activo=True)],
    }


def movement_payload(movimiento):
    created_at = timezone.localtime(movimiento.creado)
    return {
        "type": movimiento.get_tipo_display(),
        "description": movimiento.descripcion,
        "status": movimiento.get_estado_equipo_display() if movimiento.estado_equipo else "Sin estado",
        "location": movimiento.ubicacion.nombre if movimiento.ubicacion else "Sin ubicacion",
        "physicalLocation": movimiento.ubicacion_fisica.nombre if movimiento.ubicacion_fisica else "Sin registro",
        "assignedTo": movimiento.asignado_a or "Sin asignar",
        "performedBy": movimiento.realizado_por_nombre or (movimiento.realizado_por.nombre if movimiento.realizado_por else "Sin registro"),
        "date": created_at.strftime("%d/%m/%Y %H:%M"),
        "createdAt": created_at.isoformat(),
    }


def safekeeping_payload(resguardo):
    return {
        "id": resguardo.id,
        "assignedTo": resguardo.asignado_a,
        "location": resguardo.ubicacion_nombre,
        "physicalLocation": resguardo.ubicacion_fisica_nombre or "Sin registro",
        "assignmentDate": resguardo.fecha_asignacion.strftime("%d/%m/%Y"),
        "generatedAt": timezone.localtime(resguardo.generado).strftime("%d/%m/%Y %H:%M"),
        "generatedBy": resguardo.generado_por_nombre or "Sin registro",
        "signed": bool(resguardo.archivo_firmado),
        "uploadedAt": timezone.localtime(resguardo.cargado).strftime("%d/%m/%Y %H:%M") if resguardo.cargado else "",
        "uploadedBy": resguardo.cargado_por_nombre or "",
        "downloadUrl": f"/resguardos/{resguardo.id}/firmado/" if resguardo.archivo_firmado else "",
    }


def location_payload(ubicacion):
    equipment_count = ubicacion.equipo_set.count()
    return {
        "id": ubicacion.id,
        "name": ubicacion.nombre,
        "type": ubicacion.tipo,
        "manager": ubicacion.responsable_nombre,
        "managerRole": ubicacion.responsable_cargo,
        "equipmentCount": equipment_count,
        "active": ubicacion.activa,
    }


def physical_location_payload(ubicacion_fisica):
    return {
        "id": ubicacion_fisica.id,
        "locationId": ubicacion_fisica.ubicacion_id,
        "name": ubicacion_fisica.nombre,
        "active": ubicacion_fisica.activa,
    }


def get_or_create_physical_location(ubicacion, name):
    normalized_name = (name or "Santa Lucía").strip() or "Santa Lucía"
    existing = UbicacionFisica.objects.filter(
        ubicacion=ubicacion,
        nombre__iexact=normalized_name,
    ).first()
    if existing:
        if not existing.activa:
            existing.activa = True
            existing.save(update_fields=["activa"])
        return existing
    return UbicacionFisica.objects.create(ubicacion=ubicacion, nombre=normalized_name)


def responsible_payload(responsable):
    return {
        "id": responsable.id,
        "name": responsable.nombre,
        "username": responsable.usuario,
        "role": responsable.rol,
        "roleLabel": responsable.get_rol_display(),
        "locationId": responsable.ubicacion_id,
        "locationName": responsable.ubicacion.nombre if responsable.ubicacion else "Todas las ubicaciones",
        "active": responsable.activo,
    }


def equipment_type_request_payload(solicitud):
    return {
        "id": solicitud.id,
        "nombre": solicitud.nombre,
        "solicitante": solicitud.solicitante.username if solicitud.solicitante else "Usuario eliminado",
        "estado": solicitud.estado,
        "estadoLabel": solicitud.get_estado_display(),
    }


def equipment_brand_request_payload(solicitud):
    return {
        "id": solicitud.id,
        "nombre": solicitud.nombre,
        "solicitante": solicitud.solicitante.username if solicitud.solicitante else "Usuario eliminado",
        "estado": solicitud.estado,
        "estadoLabel": solicitud.get_estado_display(),
    }


# Identidad, alcance y permisos de la sesion activa.
def current_role_value(request):
    if request.user.is_staff:
        return "admin"

    responsable = Responsable.objects.filter(user=request.user, activo=True).select_related("ubicacion").first()
    if not responsable:
        return "admin"

    if responsable.rol == Responsable.Rol.MESA_TIC:
        return "tic"

    if responsable.ubicacion_id:
        return f"location:{responsable.ubicacion_id}"

    return "admin"


def current_responsable(request):
    return Responsable.objects.filter(user=request.user, activo=True).select_related("ubicacion").first()


def request_ip(request):
    return request.META.get("REMOTE_ADDR") or None


def movement_actor_data(request, responsable=None):
    responsable = responsable or current_responsable(request)
    return {
        "realizado_por": responsable,
        "realizado_por_usuario": request.user,
        "realizado_por_nombre": responsable.nombre if responsable else request.user.get_full_name() or request.user.username,
        "realizado_por_rol": responsable.get_rol_display() if responsable else "Administrador" if request.user.is_staff else "Usuario",
        "direccion_ip": request_ip(request),
        "agente_usuario": request.META.get("HTTP_USER_AGENT", ""),
    }


def has_full_inventory_access(request):
    responsable = current_responsable(request)
    return request.user.is_staff or (responsable and responsable.rol == Responsable.Rol.MESA_TIC)


def can_manage_equipment(request, equipo):
    if has_full_inventory_access(request):
        return True

    responsable = current_responsable(request)
    return bool(responsable and responsable.ubicacion_id == equipo.ubicacion_id)


# Base comun para la consulta y exportacion de reportes.
def report_equipment_queryset(request):
    queryset = Equipo.objects.select_related(
        "tipo_equipo",
        "marca_equipo",
        "ubicacion",
        "ubicacion_fisica",
    )
    if has_full_inventory_access(request):
        return queryset

    responsable = current_responsable(request)
    if not responsable or not responsable.ubicacion_id:
        return queryset.none()
    return queryset.filter(ubicacion_id=responsable.ubicacion_id)


def apply_report_location_filters(request, queryset):
    location_filter = request.GET.get("ubicacion", "").strip()
    physical_location_name = request.GET.get("ubicacion_fisica", "").strip()

    if not has_full_inventory_access(request):
        responsable = current_responsable(request)
        location_filter = str(responsable.ubicacion_id) if responsable and responsable.ubicacion_id else ""

    if location_filter.startswith("type:"):
        location_type = location_filter.removeprefix("type:")
        if location_type in {Ubicacion.Tipo.FRENTE, Ubicacion.Tipo.MESA}:
            queryset = queryset.filter(ubicacion__tipo=location_type)
    elif location_filter:
        queryset = queryset.filter(ubicacion_id=location_filter)
    if physical_location_name:
        queryset = queryset.filter(ubicacion_fisica__nombre__iexact=physical_location_name)
    return queryset


def apply_report_detail_filters(request, queryset):
    equipment_type = request.GET.get("tipo_equipo", "").strip()
    brand = request.GET.get("marca", "").strip()
    status_label = request.GET.get("estado", "").strip()
    assignment = request.GET.get("asignacion", "").strip()

    if equipment_type:
        queryset = queryset.filter(tipo__iexact=equipment_type)
    if brand == "Sin marca":
        queryset = queryset.filter(marca="")
    elif brand:
        queryset = queryset.filter(marca__iexact=brand)

    state_map = {label.lower(): value for value, label in Equipo.Estado.choices}
    if status_label:
        status = state_map.get(status_label.lower())
        queryset = queryset.filter(estado=status) if status else queryset.none()

    if assignment == "assigned":
        queryset = queryset.exclude(asignado_a="")
    elif assignment == "unassigned":
        queryset = queryset.filter(asignado_a="")
    return queryset


def style_report_worksheet(worksheet, widths):
    header_fill = PatternFill("solid", fgColor="166534")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def report_workbook_response(workbook, report_name):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    date_suffix = timezone.localdate().strftime("%Y%m%d")
    return FileResponse(
        output,
        as_attachment=True,
        filename=f"SICI-{report_name}-{date_suffix}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Lectura y validacion de valores recibidos desde archivos y formularios.
def parse_date_field(value):
    value = (value or "").strip()
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def validate_invoice_file(uploaded_file):
    if not uploaded_file:
        return None

    allowed_types = {"application/pdf", "image/jpeg", "image/png", "image/webp"}
    allowed_extensions = {".pdf", ".jpg", ".jpeg", ".png", ".webp"}
    suffix = uploaded_file.name.lower().rsplit(".", 1)
    extension = f".{suffix[-1]}" if len(suffix) > 1 else ""

    if uploaded_file.content_type not in allowed_types or extension not in allowed_extensions:
        return "La factura debe ser PDF o imagen JPG, PNG o WEBP."

    return None


def excel_date(value):
    if value in {None, ""}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value

    text = str(value).strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError


def import_rows_from_file(uploaded_file):
    extension = uploaded_file.name.lower().rsplit(".", 1)[-1]
    if extension == "xlsx":
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook["Equipos"] if "Equipos" in workbook.sheetnames else workbook.active
        values = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values, [])]
        return headers, list(values)

    if extension == "csv":
        wrapper = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig", newline="")
        reader = csv.reader(wrapper)
        headers = [value.strip() for value in next(reader, [])]
        return headers, list(reader)

    raise ValueError("Formato no soportado.")


# Catalogos, listas desplegables y archivos de apoyo para Excel.
def equipment_catalog_data(request):
    responsable = current_responsable(request)
    scoped_location = None if has_full_inventory_access(request) else responsable.ubicacion if responsable else None
    available_locations = (
        [scoped_location]
        if scoped_location
        else list(Ubicacion.objects.filter(activa=True).order_by("nombre"))
    )
    return available_locations, {
        "Tipos": list(TipoEquipo.objects.filter(activo=True).values_list("nombre", flat=True)),
        "Marcas": list(MarcaEquipo.objects.filter(activo=True).values_list("nombre", flat=True)),
        "F.F.O.O. o Mesas": [location.nombre for location in available_locations],
        "Ubicaciones fisicas": list(
            UbicacionFisica.objects.filter(
                ubicacion__in=available_locations,
                activa=True,
            )
            .values_list("nombre", flat=True)
            .distinct()
        ) or ["Santa Lucía"],
        "Estados": [label for _, label in Equipo.Estado.choices],
        "RAM": sorted(RAM_OPTIONS, key=lambda value: int(value.split()[0])),
        "Almacenamiento": sorted(STORAGE_OPTIONS),
    }


def populate_catalog_sheet(catalogs, catalog_data):
    catalogs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(catalog_data))
    note_cell = catalogs.cell(
        row=1,
        column=1,
        value=(
            "Utiliza exactamente las opciones de este catalogo. Si necesitas una marca, tipo de equipo, "
            "memoria RAM o almacenamiento que no aparece, ponte en contacto con la Mesa TIC."
        ),
    )
    note_cell.font = Font(bold=True, color="7A3E00")
    note_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    catalogs.row_dimensions[1].height = 42

    catalog_columns = {}
    for column, (title, values) in enumerate(catalog_data.items(), 1):
        title_cell = catalogs.cell(row=3, column=column, value=title)
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="176B45")
        title_cell.alignment = Alignment(horizontal="center")
        for row, value in enumerate(values, 4):
            catalogs.cell(row=row, column=column, value=value)
        catalogs.column_dimensions[title_cell.column_letter].width = max(20, len(title) + 4)
        catalog_columns[title] = (title_cell.column_letter, max(4, len(values) + 3))

    catalogs.freeze_panes = "A4"
    return catalog_columns


def add_equipment_validations(worksheet, catalog_columns, last_row=1001):
    validations = {
        "A": "Tipos",
        "B": "Marcas",
        "F": "RAM",
        "G": "Almacenamiento",
        "M": "F.F.O.O. o Mesas",
        "N": "Ubicaciones fisicas",
        "O": "Estados",
    }
    for target_column, catalog_name in validations.items():
        catalog_column, catalog_last_row = catalog_columns[catalog_name]
        validation = DataValidation(
            type="list",
            formula1=f"'Catalogos'!${catalog_column}$4:${catalog_column}${catalog_last_row}",
            allow_blank=target_column not in {"A", "M", "N", "O"},
        )
        worksheet.add_data_validation(validation)
        validation.add(f"{target_column}2:{target_column}{last_row}")


def rejection_reason(error):
    return re.sub(r"^Fila \d+:\s*", "", str(error)).strip()


def rejected_rows_workbook(headers, rejected_rows, catalog_data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Equipos rechazados"
    catalogs = workbook.create_sheet("Catalogos")
    output_headers = [*headers, "observaciones_rechazo"]
    worksheet.append(output_headers)

    for row, reason in rejected_rows:
        worksheet.append([row.get(header, "") for header in headers] + [reason])

    header_fill = PatternFill("solid", fgColor="A61B1B")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = [20, 18, 20, 24, 22, 16, 20, 22, 20, 20, 16, 16, 24, 22, 18, 28, 36, 60]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    catalog_columns = populate_catalog_sheet(catalogs, catalog_data)
    add_equipment_validations(worksheet, catalog_columns, max(1001, len(rejected_rows) + 1))

    buffer = BytesIO()
    workbook.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


# Descarga de plantilla y procesamiento de cargas masivas.
@login_required
def plantilla_equipos_excel(request):
    available_locations, catalog_data = equipment_catalog_data(request)
    responsable = current_responsable(request)
    scoped_location = None if has_full_inventory_access(request) else responsable.ubicacion if responsable else None
    example_location = scoped_location or next(
        (location for location in available_locations if location.nombre.strip().lower() in {"tic", "mesa tic"}),
        available_locations[0] if available_locations else None,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Equipos"
    catalogs = workbook.create_sheet("Catalogos")

    for column, header in enumerate(EQUIPMENT_IMPORT_HEADERS, 1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B45")
        cell.alignment = Alignment(horizontal="center")

    example = [
        "Laptop",
        "Dell",
        "Latitude 5440",
        "SERIE-EJEMPLO-001",
        "Intel Core i7",
        "16 GB",
        "512 GB SSD",
        "Windows 11 Pro",
        "00:1A:2B:3C:4D:5E",
        "00:1A:2B:3C:4D:5F",
        "2026-01-15",
        "2029-01-15",
        example_location.nombre if example_location else "F.F.O.O. o Mesa",
        "Santa Lucía",
        "Disponible",
        "",
        "Fila de ejemplo: eliminar antes de importar.",
    ]
    worksheet.append(example)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:Q2"

    widths = [20, 18, 20, 24, 22, 16, 20, 22, 20, 20, 16, 16, 24, 22, 18, 28, 36]
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width

    catalog_columns = populate_catalog_sheet(catalogs, catalog_data)
    add_equipment_validations(worksheet, catalog_columns)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Plantilla-importacion-SICI.xlsx")


@login_required
@require_POST
@transaction.atomic
def importar_equipos(request):
    preview = request.POST.get("preview", "").strip().lower() in {"true", "1", "si"}
    uploaded_file = request.FILES.get("archivo")
    if not uploaded_file:
        return JsonResponse({"error": "Selecciona un archivo Excel o CSV."}, status=400)
    if uploaded_file.size > 10 * 1024 * 1024:
        return JsonResponse({"error": "El archivo no debe superar 10 MB."}, status=400)

    try:
        headers, raw_rows = import_rows_from_file(uploaded_file)
    except (ValueError, KeyError, OSError):
        return JsonResponse({"error": "No fue posible leer el archivo. Usa la plantilla XLSX o un CSV UTF-8."}, status=400)

    accepted_headers = [
        EQUIPMENT_IMPORT_HEADERS,
        [*EQUIPMENT_IMPORT_HEADERS, "observaciones_rechazo"],
    ]
    if headers not in accepted_headers:
        return JsonResponse({"error": "Las columnas no coinciden con la plantilla oficial de SICI."}, status=400)
    headers = EQUIPMENT_IMPORT_HEADERS

    rows = []
    for row_number, values in enumerate(raw_rows, 2):
        row = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}
        if any(str(value or "").strip() for value in row.values()):
            rows.append((row_number, row))
    if not rows:
        return JsonResponse({"error": "El archivo no contiene equipos para importar."}, status=400)

    types = {item.nombre.lower(): item for item in TipoEquipo.objects.filter(activo=True)}
    brands = {item.nombre.lower(): item for item in MarcaEquipo.objects.filter(activo=True)}
    locations_by_name = {item.nombre.lower(): item for item in Ubicacion.objects.filter(activa=True)}
    state_map = {label.lower(): value for value, label in Equipo.Estado.choices}
    responsable = current_responsable(request)
    _, catalog_data = equipment_catalog_data(request)
    imported = []
    rejected = []
    preview_valid = []
    preview_errors = []
    type_summary = {}
    seen_serials = set()

    for row_number, row in rows:
        try:
            with transaction.atomic():
                clean = {key: str(value or "").strip() for key, value in row.items()}
                type_name = clean["tipo_equipo"]
                serial = clean["numero_serie"]
                location_name = clean["ffoo_o_mesa"]
                physical_location_name = clean["ubicacion_fisica"] or "Santa Lucía"
                state_label = clean["estado"]
                if not type_name or not serial or not location_name or not state_label:
                    raise ValueError(
                        f"Fila {row_number}: tipo_equipo, numero_serie, ffoo_o_mesa y estado son obligatorios."
                    )

                serial_key = serial.lower()
                if serial_key in seen_serials or Equipo.objects.filter(numero_serie__iexact=serial).exists():
                    raise ValueError(f"Fila {row_number}: el numero de serie {serial} ya existe o esta repetido.")

                equipment_type = types.get(type_name.lower())
                if not equipment_type:
                    raise ValueError(f"Fila {row_number}: el tipo de equipo '{type_name}' no esta registrado.")

                brand = None
                if clean["marca"]:
                    brand = brands.get(clean["marca"].lower())
                    if not brand:
                        raise ValueError(f"Fila {row_number}: la marca '{clean['marca']}' no esta registrada.")

                location = locations_by_name.get(location_name.lower())
                if not location:
                    raise ValueError(f"Fila {row_number}: la ubicacion '{location_name}' no esta registrada.")
                if not has_full_inventory_access(request) and (
                    not responsable or responsable.ubicacion_id != location.id
                ):
                    expected_location = responsable.ubicacion.nombre if responsable and responsable.ubicacion else "tu F.F.O.O. o Mesa"
                    raise PermissionError(
                        f"Fila {row_number}: F.F.O.O. o Mesa debe ser '{expected_location}'."
                    )
                physical_location = get_or_create_physical_location(location, physical_location_name)

                state = state_map.get(state_label.lower())
                if not state:
                    raise ValueError(f"Fila {row_number}: el estado '{state_label}' no es valido.")
                assigned_to = clean["asignado_a"]
                if state == Equipo.Estado.ASIGNADO and not assigned_to:
                    raise ValueError(f"Fila {row_number}: asignado_a es obligatorio para equipos asignados.")
                if state in {Equipo.Estado.DISPONIBLE, Equipo.Estado.BAJA}:
                    assigned_to = ""

                ram = clean["memoria_ram"]
                storage = clean["almacenamiento"]
                if ram and ram not in RAM_OPTIONS:
                    raise ValueError(f"Fila {row_number}: memoria RAM no valida.")
                if storage and storage not in STORAGE_OPTIONS:
                    raise ValueError(f"Fila {row_number}: almacenamiento no valido.")

                try:
                    purchase_date = excel_date(row["fecha_compra"])
                    warranty_date = excel_date(row["garantia_hasta"])
                except ValueError:
                    raise ValueError(f"Fila {row_number}: usa fechas YYYY-MM-DD o DD/MM/YYYY.")

                equipment = Equipo.objects.create(
                    tipo=equipment_type.nombre,
                    tipo_equipo=equipment_type,
                    marca_equipo=brand,
                    marca=brand.nombre if brand else "",
                    modelo=clean["modelo"],
                    numero_serie=serial,
                    procesador=clean["procesador"],
                    memoria_ram=ram,
                    almacenamiento=storage,
                    sistema_operativo=clean["sistema_operativo"],
                    mac_ethernet=clean["mac_ethernet"],
                    mac_wifi=clean["mac_wifi"],
                    fecha_compra=purchase_date,
                    garantia_hasta=warranty_date,
                    estado=state,
                    ubicacion=location,
                    ubicacion_fisica=physical_location,
                    asignado_a=assigned_to,
                    observaciones=clean["observaciones"],
                )
                movement_type = {
                    Equipo.Estado.ASIGNADO: Movimiento.Tipo.ASIGNACION,
                    Equipo.Estado.MANTENIMIENTO: Movimiento.Tipo.MANTENIMIENTO,
                    Equipo.Estado.BAJA: Movimiento.Tipo.BAJA,
                }.get(state, Movimiento.Tipo.ALTA)
                Movimiento.objects.create(
                    equipo=equipment,
                    tipo=movement_type,
                    descripcion=f"Alta mediante importacion de Excel con estado {state_label}.",
                    **movement_actor_data(request, responsable),
                    ubicacion=location,
                    ubicacion_fisica=physical_location,
                    estado_equipo=state,
                    asignado_a=assigned_to,
                )
                imported.append(equipment_payload(equipment))
                preview_valid.append(
                    {
                        "row": row_number,
                        "type": equipment_type.nombre,
                        "serial": serial,
                        "location": location.nombre,
                        "status": state_label,
                    }
                )
                type_summary[equipment_type.nombre] = type_summary.get(equipment_type.nombre, 0) + 1
                seen_serials.add(serial_key)
        except (PermissionError, ValueError, IntegrityError) as error:
            reason = rejection_reason(error)
            rejected.append((row, reason))
            preview_errors.append(
                {
                    "row": row_number,
                    "type": str(row.get("tipo_equipo") or "").strip() or "Sin tipo",
                    "serial": str(row.get("numero_serie") or "").strip() or "Sin serie",
                    "reason": reason,
                }
            )

    if preview:
        transaction.set_rollback(True)
        return JsonResponse(
            {
                "preview": True,
                "filename": uploaded_file.name,
                "totalCount": len(rows),
                "validCount": len(imported),
                "rejectedCount": len(rejected),
                "typeSummary": [
                    {"type": equipment_type, "count": count}
                    for equipment_type, count in sorted(
                        type_summary.items(),
                        key=lambda item: (-item[1], item[0].lower()),
                    )
                ],
                "validRows": preview_valid,
                "errors": preview_errors,
            }
        )

    response_data = {
        "equipment": imported,
        "count": len(imported),
        "rejectedCount": len(rejected),
    }
    if rejected:
        response_data.update(
            {
                "rejectedFile": rejected_rows_workbook(headers, rejected, catalog_data),
                "rejectedFilename": f"Equipos-rechazados-SICI-{timezone.localdate():%Y%m%d}.xlsx",
            }
        )

    return JsonResponse(response_data, status=201 if imported else 200)


# Contexto inicial de la aplicacion y datos de la sesion.
def session_user_payload(request):
    responsable = Responsable.objects.filter(user=request.user, activo=True).select_related("ubicacion").first()
    if responsable:
        return {
            "name": responsable.nombre,
            "username": responsable.usuario,
            "role": responsable.get_rol_display(),
            "scope": responsable.ubicacion.nombre if responsable.ubicacion else "Todas las ubicaciones",
        }

    return {
        "name": request.user.get_full_name() or request.user.username,
        "username": request.user.username,
        "role": "Administrador" if request.user.is_staff else "Usuario",
        "scope": "Todas las ubicaciones" if request.user.is_staff else "Sin ubicacion asignada",
    }


@login_required
def home(request):
    equipos = Equipo.objects.select_related("ubicacion", "ubicacion_fisica").prefetch_related("movimientos", "resguardos").all()
    ubicaciones = Ubicacion.objects.filter(activa=True).order_by("nombre")
    ubicaciones_fisicas = UbicacionFisica.objects.filter(
        ubicacion__in=ubicaciones,
        activa=True,
    ).select_related("ubicacion")
    responsables = Responsable.objects.select_related("ubicacion", "user").order_by("nombre")
    tipos_equipo = TipoEquipo.objects.filter(activo=True).order_by("nombre")
    marcas_equipo = MarcaEquipo.objects.filter(activo=True).order_by("nombre")
    solicitudes_tipo = SolicitudTipoEquipo.objects.filter(estado=SolicitudTipoEquipo.Estado.PENDIENTE)
    solicitudes_marca = SolicitudMarcaEquipo.objects.filter(estado=SolicitudMarcaEquipo.Estado.PENDIENTE)
    context = {
        "equipment_data": [equipment_payload(equipo) for equipo in equipos],
        "location_data": [location_payload(ubicacion) for ubicacion in ubicaciones],
        "physical_location_data": [
            physical_location_payload(ubicacion_fisica)
            for ubicacion_fisica in ubicaciones_fisicas
        ],
        "responsible_data": [responsible_payload(responsable) for responsable in responsables],
        "equipment_type_data": list(tipos_equipo.values("id", "nombre")),
        "equipment_brand_data": list(marcas_equipo.values("id", "nombre")),
        "equipment_type_request_data": [
            equipment_type_request_payload(solicitud) for solicitud in solicitudes_tipo
        ],
        "equipment_brand_request_data": [
            equipment_brand_request_payload(solicitud) for solicitud in solicitudes_marca
        ],
        "current_role_value": current_role_value(request),
        "session_user": session_user_payload(request),
        "can_manage_equipment_types": request.user.is_staff,
    }
    return render(request, "inventario/index.html", context)


# Generacion de reportes operativos en formato Excel.
@login_required
def descargar_reporte(request):
    report_type = request.GET.get("tipo", "inventario").strip()
    allowed_reports = {"inventario", "garantias", "mantenimiento", "bajas", "movimientos"}
    if report_type not in allowed_reports:
        return HttpResponse("Tipo de reporte no valido.", status=400)

    equipment_queryset = apply_report_detail_filters(
        request,
        apply_report_location_filters(request, report_equipment_queryset(request)),
    )
    workbook = Workbook()
    worksheet = workbook.active

    if report_type == "movimientos":
        worksheet.title = "Movimientos"
        worksheet.append(
            [
                "Fecha y hora",
                "Movimiento",
                "Numero de serie",
                "Equipo",
                "F.F.O.O. o Mesa",
                "Ubicacion fisica",
                "Estado",
                "Asignado a",
                "Realizado por",
                "Rol",
                "Direccion IP",
                "Descripcion",
            ]
        )
        movements = (
            Movimiento.objects.filter(equipo__in=equipment_queryset)
            .select_related("equipo", "ubicacion", "ubicacion_fisica")
            .order_by("-creado")
        )
        for movement in movements:
            created_at = timezone.localtime(movement.creado)
            worksheet.append(
                [
                    created_at.replace(tzinfo=None),
                    movement.get_tipo_display(),
                    movement.equipo.numero_serie,
                    str(movement.equipo),
                    movement.ubicacion.nombre if movement.ubicacion else "",
                    movement.ubicacion_fisica.nombre if movement.ubicacion_fisica else "",
                    movement.get_estado_equipo_display() if movement.estado_equipo else "",
                    movement.asignado_a,
                    movement.realizado_por_nombre,
                    movement.realizado_por_rol,
                    str(movement.direccion_ip or ""),
                    movement.descripcion,
                ]
            )
        style_report_worksheet(worksheet, [20, 18, 24, 32, 25, 24, 18, 28, 25, 22, 18, 55])
        for cell in worksheet["A"][1:]:
            cell.number_format = "dd/mm/yyyy hh:mm"
        return report_workbook_response(workbook, "movimientos")

    today = timezone.localdate()
    report_labels = {
        "inventario": ("Inventario", "inventario"),
        "garantias": ("Garantias", "garantias"),
        "mantenimiento": ("Mantenimiento", "mantenimiento"),
        "bajas": ("Bajas", "bajas"),
    }
    worksheet.title = report_labels[report_type][0]

    if report_type == "garantias":
        warranty_filter = request.GET.get("garantia", request.GET.get("dias", "90")).strip()
        if warranty_filter == "none":
            equipment_queryset = equipment_queryset.filter(garantia_hasta__isnull=True)
        elif warranty_filter == "expired":
            equipment_queryset = equipment_queryset.filter(garantia_hasta__lt=today)
        elif warranty_filter == "all":
            equipment_queryset = equipment_queryset.filter(garantia_hasta__isnull=False)
        else:
            try:
                warranty_days = min(max(int(warranty_filter), 1), 730)
            except ValueError:
                warranty_days = 90
            equipment_queryset = equipment_queryset.filter(
                garantia_hasta__isnull=False,
                garantia_hasta__gte=today,
                garantia_hasta__lte=today + timedelta(days=warranty_days),
            )
        equipment_queryset = equipment_queryset.order_by("garantia_hasta", "ubicacion__nombre")
    elif report_type == "mantenimiento":
        equipment_queryset = equipment_queryset.filter(estado=Equipo.Estado.MANTENIMIENTO)
    elif report_type == "bajas":
        equipment_queryset = equipment_queryset.filter(estado=Equipo.Estado.BAJA)

    worksheet.append(
        [
            "Tipo",
            "Marca",
            "Modelo",
            "Numero de serie",
            "F.F.O.O. o Mesa",
            "Ubicacion fisica",
            "Estado",
            "Asignado a",
            "Fecha de compra",
            "Garantia hasta",
            "Procesador",
            "RAM",
            "Almacenamiento",
            "Sistema operativo",
            "MAC Ethernet",
            "MAC WiFi",
            "Observaciones",
        ]
    )
    for equipment_item in equipment_queryset.order_by("ubicacion__nombre", "tipo", "numero_serie"):
        worksheet.append(
            [
                equipment_item.tipo,
                equipment_item.marca,
                equipment_item.modelo,
                equipment_item.numero_serie,
                equipment_item.ubicacion.nombre,
                equipment_item.ubicacion_fisica.nombre if equipment_item.ubicacion_fisica else "",
                equipment_item.get_estado_display(),
                equipment_item.asignado_a,
                equipment_item.fecha_compra,
                equipment_item.garantia_hasta,
                equipment_item.procesador,
                equipment_item.memoria_ram,
                equipment_item.almacenamiento,
                equipment_item.sistema_operativo,
                equipment_item.mac_ethernet,
                equipment_item.mac_wifi,
                equipment_item.observaciones,
            ]
        )
    style_report_worksheet(
        worksheet,
        [18, 18, 20, 25, 25, 24, 18, 28, 16, 16, 24, 14, 20, 24, 20, 20, 45],
    )
    for column in ("I", "J"):
        for cell in worksheet[column][1:]:
            cell.number_format = "dd/mm/yyyy"
    return report_workbook_response(workbook, report_labels[report_type][1])


# Validaciones compartidas por los modulos administrativos.
def admin_required_json(request):
    if request.user.is_staff:
        return None
    return JsonResponse({"error": "Solo el administrador puede realizar esta accion."}, status=403)


def parse_bool(value):
    return str(value).lower() in {"true", "1", "activo", "on"}


def normalize_role(value):
    role = (value or "").strip()
    allowed = {choice.value for choice in Responsable.Rol}
    return role if role in allowed else ""


def get_responsable_form_data(request, require_password=False):
    nombre = request.POST.get("nombre", "").strip()
    usuario = request.POST.get("usuario", "").strip()
    password = request.POST.get("password", "")
    rol = normalize_role(request.POST.get("rol"))
    ubicacion_id = request.POST.get("ubicacion") or None
    activo = parse_bool(request.POST.get("activo", "true"))

    if not nombre or not usuario or not rol:
        return None, "Nombre, usuario y rol son obligatorios."

    if require_password and not password:
        return None, "La contrasena inicial es obligatoria."

    ubicacion = None
    if rol in {Responsable.Rol.FRENTE, Responsable.Rol.MESA}:
        if not ubicacion_id:
            return None, "Selecciona el frente o mesa que administrara."
        ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)

    return {
        "nombre": nombre,
        "usuario": usuario,
        "password": password,
        "rol": rol,
        "ubicacion": ubicacion,
        "activo": activo,
    }, None


# Consulta y navegacion mediante codigos QR.
@login_required
def equipo_qr(request, numero_serie):
    equipo = get_object_or_404(Equipo, numero_serie=numero_serie)
    qr_target = request.build_absolute_uri(f"/q/{quote(equipo.numero_serie, safe='')}/")
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=16,
        border=4,
    )
    qr.add_data(qr_target)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    response = HttpResponse(buffer.getvalue(), content_type="image/png")
    response["Content-Disposition"] = f'inline; filename="SICI-{equipo.numero_serie}.png"'
    return response


@login_required
def equipo_qr_link(request, numero_serie):
    equipo = get_object_or_404(Equipo, numero_serie=numero_serie)
    return redirect(f"/?{urlencode({'qr': equipo.numero_serie})}")


# Administracion de Frentes de Obra y Mesas.
@login_required
@require_POST
def crear_ubicacion(request):
    error = admin_required_json(request)
    if error:
        return error

    tipo = request.POST.get("tipo")
    nombre = request.POST.get("nombre", "").strip()
    responsable_nombre = request.POST.get("responsable_nombre", "").strip()
    activa = parse_bool(request.POST.get("activa"))

    if tipo not in {Ubicacion.Tipo.FRENTE, Ubicacion.Tipo.MESA}:
        return JsonResponse({"error": "Tipo de ubicacion invalido."}, status=400)

    if not nombre or not responsable_nombre:
        return JsonResponse({"error": "Nombre y responsable son obligatorios."}, status=400)

    responsable_cargo = "I.R.O." if tipo == Ubicacion.Tipo.FRENTE else "Jefe de Mesa"

    ubicacion, created = Ubicacion.objects.get_or_create(
        nombre=nombre,
        defaults={
            "tipo": tipo,
            "responsable_nombre": responsable_nombre,
            "responsable_cargo": responsable_cargo,
            "activa": activa,
        },
    )

    if not created:
        return JsonResponse({"error": "Ya existe una ubicacion con ese nombre."}, status=400)

    physical_location, _ = UbicacionFisica.objects.get_or_create(
        ubicacion=ubicacion,
        nombre="Santa Lucía",
    )
    return JsonResponse(
        {
            "location": location_payload(ubicacion),
            "physicalLocation": physical_location_payload(physical_location),
        },
        status=201,
    )


@login_required
@require_POST
def editar_ubicacion(request, ubicacion_id):
    error = admin_required_json(request)
    if error:
        return error

    ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)
    tipo = request.POST.get("tipo")
    nombre = request.POST.get("nombre", "").strip()
    responsable_nombre = request.POST.get("responsable_nombre", "").strip()

    if tipo not in {Ubicacion.Tipo.FRENTE, Ubicacion.Tipo.MESA}:
        return JsonResponse({"error": "Tipo de ubicacion invalido."}, status=400)

    if not nombre or not responsable_nombre:
        return JsonResponse({"error": "Nombre y responsable son obligatorios."}, status=400)

    if Ubicacion.objects.exclude(pk=ubicacion.pk).filter(nombre__iexact=nombre).exists():
        return JsonResponse({"error": "Ya existe una ubicacion con ese nombre."}, status=400)

    ubicacion.tipo = tipo
    ubicacion.nombre = nombre
    ubicacion.responsable_nombre = responsable_nombre
    ubicacion.responsable_cargo = "I.R.O." if tipo == Ubicacion.Tipo.FRENTE else "Jefe de Mesa"
    ubicacion.activa = parse_bool(request.POST.get("activa"))
    ubicacion.save()
    return JsonResponse({"location": location_payload(ubicacion)})


@login_required
@require_POST
def eliminar_ubicacion(request, ubicacion_id):
    error = admin_required_json(request)
    if error:
        return error

    ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)
    if ubicacion.equipo_set.exists() or ubicacion.responsable_set.exists():
        return JsonResponse(
            {"error": "No se puede eliminar porque tiene equipos o responsables asociados."},
            status=400,
        )

    ubicacion.delete()
    return JsonResponse({"deleted": ubicacion_id})


# Administracion de responsables, cuentas y contrasenas.
@login_required
@require_POST
def crear_responsable(request):
    error = admin_required_json(request)
    if error:
        return error

    data, error_message = get_responsable_form_data(request, require_password=True)
    if error_message:
        return JsonResponse({"error": error_message}, status=400)

    try:
        with transaction.atomic():
            user = User.objects.create_user(
                username=data["usuario"],
                password=data["password"],
                first_name=data["nombre"],
                is_staff=data["rol"] == Responsable.Rol.ADMIN,
                is_active=data["activo"],
            )
            responsable = Responsable.objects.create(
                user=user,
                nombre=data["nombre"],
                usuario=data["usuario"],
                rol=data["rol"],
                ubicacion=data["ubicacion"],
                activo=data["activo"],
            )
    except IntegrityError:
        return JsonResponse({"error": "Ya existe un usuario con ese nombre."}, status=400)

    return JsonResponse({"responsible": responsible_payload(responsable)}, status=201)


@login_required
@require_POST
def editar_responsable(request, responsable_id):
    error = admin_required_json(request)
    if error:
        return error

    responsable = get_object_or_404(Responsable.objects.select_related("user"), pk=responsable_id)
    data, error_message = get_responsable_form_data(request)
    if error_message:
        return JsonResponse({"error": error_message}, status=400)

    if Responsable.objects.exclude(pk=responsable.pk).filter(usuario__iexact=data["usuario"]).exists():
        return JsonResponse({"error": "Ya existe un responsable con ese usuario."}, status=400)

    if User.objects.exclude(pk=responsable.user_id).filter(username__iexact=data["usuario"]).exists():
        return JsonResponse({"error": "Ya existe un usuario con ese nombre."}, status=400)

    with transaction.atomic():
        responsable.nombre = data["nombre"]
        responsable.usuario = data["usuario"]
        responsable.rol = data["rol"]
        responsable.ubicacion = data["ubicacion"]
        responsable.activo = data["activo"]
        responsable.save()

        if responsable.user:
            responsable.user.username = data["usuario"]
            responsable.user.first_name = data["nombre"]
            responsable.user.is_staff = data["rol"] == Responsable.Rol.ADMIN
            responsable.user.is_active = data["activo"]
            responsable.user.save()

    return JsonResponse({"responsible": responsible_payload(responsable)})


@login_required
@require_POST
def cambiar_password_responsable(request, responsable_id):
    error = admin_required_json(request)
    if error:
        return error

    responsable = get_object_or_404(Responsable.objects.select_related("user"), pk=responsable_id)
    password = request.POST.get("password", "")
    if not password:
        return JsonResponse({"error": "La nueva contrasena es obligatoria."}, status=400)

    if not responsable.user:
        return JsonResponse({"error": "Este responsable no tiene usuario ligado."}, status=400)

    responsable.user.set_password(password)
    responsable.user.save()
    return JsonResponse({"ok": True})


@login_required
@require_POST
def eliminar_responsable(request, responsable_id):
    error = admin_required_json(request)
    if error:
        return error

    responsable = get_object_or_404(Responsable.objects.select_related("user"), pk=responsable_id)
    if responsable.user_id == request.user.id:
        return JsonResponse({"error": "No puedes eliminar tu propio usuario mientras estas conectado."}, status=400)

    user = responsable.user
    responsable.delete()
    if user:
        user.delete()
    return JsonResponse({"deleted": responsable_id})


# Alta, actualizacion de estado y eliminacion auditada de equipos.
@login_required
@require_POST
def crear_equipo(request):
    tipo_nombre = request.POST.get("tipo_equipo", "").strip()
    numero_serie = request.POST.get("numero_serie", "").strip()
    ubicacion_id = request.POST.get("ubicacion", "").strip()
    estado_display = request.POST.get("estado", "").strip()
    asignado_a = request.POST.get("asignado_a", "").strip()

    if not tipo_nombre or not numero_serie or not ubicacion_id or not estado_display:
        return JsonResponse({"error": "Tipo de equipo, numero de serie, ubicacion y estado son obligatorios."}, status=400)

    estado_map = {label: value for value, label in Equipo.Estado.choices}
    estado = estado_map.get(estado_display)
    if not estado:
        return JsonResponse({"error": "Estado invalido."}, status=400)

    if estado == Equipo.Estado.ASIGNADO and not asignado_a:
        return JsonResponse({"error": "Indica a quien esta asignado el equipo."}, status=400)

    if estado == Equipo.Estado.DISPONIBLE:
        asignado_a = ""

    tipo_equipo = TipoEquipo.objects.filter(nombre__iexact=tipo_nombre, activo=True).first()
    if not tipo_equipo:
        return JsonResponse({"error": "Selecciona un tipo de equipo registrado."}, status=400)

    marca_nombre = request.POST.get("marca", "").strip()
    marca_equipo = None
    if marca_nombre:
        marca_equipo = MarcaEquipo.objects.filter(nombre__iexact=marca_nombre, activo=True).first()
        if not marca_equipo:
            return JsonResponse({"error": "Selecciona una marca registrada o solicita al administrador crearla."}, status=400)
        marca_nombre = marca_equipo.nombre

    ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id, activa=True)
    responsable = current_responsable(request)
    if not has_full_inventory_access(request):
        if not responsable or responsable.ubicacion_id != ubicacion.id:
            return JsonResponse({"error": "Solo puedes registrar equipos en tu frente o mesa."}, status=403)
    ubicacion_fisica = get_or_create_physical_location(
        ubicacion,
        request.POST.get("ubicacion_fisica", ""),
    )

    if Equipo.objects.filter(numero_serie__iexact=numero_serie).exists():
        return JsonResponse({"error": "Ya existe un equipo con ese numero de serie."}, status=400)

    memoria_ram = request.POST.get("memoria_ram", "").strip()
    almacenamiento = request.POST.get("almacenamiento", "").strip()
    if memoria_ram and memoria_ram not in RAM_OPTIONS:
        return JsonResponse({"error": "Selecciona una memoria RAM valida del catalogo."}, status=400)
    if almacenamiento and almacenamiento not in STORAGE_OPTIONS:
        return JsonResponse({"error": "Selecciona un almacenamiento valido del catalogo."}, status=400)

    invoice_error = validate_invoice_file(request.FILES.get("factura"))
    if invoice_error:
        return JsonResponse({"error": invoice_error}, status=400)

    try:
        fecha_compra = parse_date_field(request.POST.get("fecha_compra"))
        garantia_hasta = parse_date_field(request.POST.get("garantia_hasta"))
    except ValueError:
        return JsonResponse({"error": "Formato de fecha invalido."}, status=400)

    with transaction.atomic():
        equipo = Equipo.objects.create(
            tipo=tipo_equipo.nombre,
            tipo_equipo=tipo_equipo,
            marca_equipo=marca_equipo,
            marca=marca_nombre,
            modelo=request.POST.get("modelo", "").strip(),
            numero_serie=numero_serie,
            procesador=request.POST.get("procesador", "").strip(),
            memoria_ram=memoria_ram,
            almacenamiento=almacenamiento,
            sistema_operativo=request.POST.get("sistema_operativo", "").strip(),
            mac_ethernet=request.POST.get("mac_ethernet", "").strip(),
            mac_wifi=request.POST.get("mac_wifi", "").strip(),
            fecha_compra=fecha_compra,
            garantia_hasta=garantia_hasta,
            factura=request.FILES.get("factura"),
            estado=estado,
            ubicacion=ubicacion,
            ubicacion_fisica=ubicacion_fisica,
            asignado_a=asignado_a,
            observaciones=request.POST.get("observaciones", "").strip(),
        )

        if estado == Equipo.Estado.ASIGNADO:
            movimiento_tipo = Movimiento.Tipo.ASIGNACION
            descripcion = f"Alta de equipo y asignacion inicial a {asignado_a}."
        elif estado == Equipo.Estado.MANTENIMIENTO:
            movimiento_tipo = Movimiento.Tipo.MANTENIMIENTO
            descripcion = "Alta de equipo en mantenimiento."
        elif estado == Equipo.Estado.BAJA:
            movimiento_tipo = Movimiento.Tipo.BAJA
            descripcion = "Alta de equipo dado de baja."
        else:
            movimiento_tipo = Movimiento.Tipo.ALTA
            descripcion = "Alta de equipo disponible."

        Movimiento.objects.create(
            equipo=equipo,
            tipo=movimiento_tipo,
            descripcion=descripcion,
            **movement_actor_data(request, responsable),
            ubicacion=ubicacion,
            ubicacion_fisica=ubicacion_fisica,
            estado_equipo=estado,
            asignado_a=asignado_a,
        )

    return JsonResponse({"equipment": equipment_payload(equipo)}, status=201)


@login_required
@require_POST
def cambiar_estado_equipo(request, numero_serie):
    equipo = get_object_or_404(Equipo.objects.select_related("ubicacion"), numero_serie=numero_serie)
    if not can_manage_equipment(request, equipo):
        return JsonResponse({"error": "Solo puedes modificar equipos de tu frente o mesa."}, status=403)

    estado_display = request.POST.get("estado", "").strip()
    motivo = request.POST.get("motivo", "").strip()
    estado_map = {label: value for value, label in Equipo.Estado.choices}
    estado = estado_map.get(estado_display)
    if not estado:
        return JsonResponse({"error": "Estado invalido."}, status=400)

    asignado_a = request.POST.get("asignado_a", "").strip()
    if estado == Equipo.Estado.ASIGNADO and not asignado_a:
        return JsonResponse({"error": "Indica a quien esta asignado el equipo."}, status=400)

    if estado == Equipo.Estado.MANTENIMIENTO and not motivo:
        return JsonResponse({"error": "Indica la razon por la que el equipo pasa a mantenimiento."}, status=400)

    if estado in {Equipo.Estado.DISPONIBLE, Equipo.Estado.BAJA}:
        asignado_a = ""

    ubicacion = equipo.ubicacion
    ubicacion_id = request.POST.get("ubicacion", "").strip()
    if ubicacion_id:
        ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id, activa=True)
        if not has_full_inventory_access(request) and ubicacion.id != equipo.ubicacion_id:
            return JsonResponse({"error": "No puedes mover equipos fuera de tu frente o mesa."}, status=403)
    ubicacion_fisica = get_or_create_physical_location(
        ubicacion,
        request.POST.get("ubicacion_fisica", "")
        or (equipo.ubicacion_fisica.nombre if equipo.ubicacion_fisica else "Santa Lucía"),
    )

    responsable = current_responsable(request)
    tipo_movimiento = Movimiento.Tipo.CAMBIO_ESTADO
    descripcion = f"Cambio de estado a {estado_display}."
    if estado == Equipo.Estado.ASIGNADO:
        tipo_movimiento = Movimiento.Tipo.ASIGNACION
        descripcion = f"Asignacion a {asignado_a}."
    elif estado == Equipo.Estado.MANTENIMIENTO:
        tipo_movimiento = Movimiento.Tipo.MANTENIMIENTO
        descripcion = f"Equipo marcado en mantenimiento. Razon: {motivo}"
    elif estado == Equipo.Estado.BAJA:
        tipo_movimiento = Movimiento.Tipo.BAJA
        descripcion = f"Equipo dado de baja.{f' Motivo: {motivo}' if motivo else ''}"
    elif motivo:
        descripcion = f"{descripcion} Motivo: {motivo}"

    equipo.estado = estado
    equipo.asignado_a = asignado_a
    equipo.ubicacion = ubicacion
    equipo.ubicacion_fisica = ubicacion_fisica
    equipo.save(update_fields=["estado", "asignado_a", "ubicacion", "ubicacion_fisica", "actualizado"])

    Movimiento.objects.create(
        equipo=equipo,
        tipo=tipo_movimiento,
        descripcion=descripcion,
        **movement_actor_data(request, responsable),
        ubicacion=ubicacion,
        ubicacion_fisica=ubicacion_fisica,
        estado_equipo=estado,
        asignado_a=asignado_a,
    )

    return JsonResponse({"equipment": equipment_payload(equipo)})


@login_required
@require_POST
def eliminar_equipo(request, numero_serie):
    equipo = get_object_or_404(Equipo.objects.select_related("ubicacion"), numero_serie=numero_serie)
    if not has_full_inventory_access(request):
        return JsonResponse({"error": "Solo el administrador o Mesa TIC pueden eliminar equipos."}, status=403)

    responsable = current_responsable(request)
    deleted_serial = equipo.numero_serie
    movimientos = [
        movement_payload(movimiento)
        for movimiento in equipo.movimientos.select_related("ubicacion", "realizado_por").all()
    ]
    snapshot = equipment_payload(equipo)
    snapshot["history"] = movimientos

    EquipoEliminado.objects.create(
        numero_serie=equipo.numero_serie,
        descripcion_equipo=str(equipo),
        eliminado_por_usuario=request.user,
        eliminado_por_responsable=responsable,
        eliminado_por_nombre=responsable.nombre if responsable else request.user.get_full_name() or request.user.username,
        eliminado_por_rol=responsable.get_rol_display() if responsable else "Administrador" if request.user.is_staff else "Usuario",
        direccion_ip=request_ip(request),
        agente_usuario=request.META.get("HTTP_USER_AGENT", ""),
        ubicacion_nombre=equipo.ubicacion.nombre,
        estado=equipo.estado,
        asignado_a=equipo.asignado_a,
        snapshot=snapshot,
    )
    equipo.delete()
    return JsonResponse({"deleted": deleted_serial})


# Preparacion del documento de resguardo.
def assignment_date_for_equipment(equipo):
    assignment = equipo.movimientos.filter(tipo=Movimiento.Tipo.ASIGNACION).first()
    return timezone.localtime(assignment.creado).date() if assignment else timezone.localdate(equipo.actualizado)


def equipment_description_for_pdf(equipo):
    parts = [
        f"{equipo.tipo}: {equipo.marca} {equipo.modelo}".strip(),
        f"Numero de serie: {equipo.numero_serie}",
    ]
    if equipo.procesador:
        parts.append(f"Procesador: {equipo.procesador}")
    if equipo.memoria_ram:
        parts.append(f"RAM: {equipo.memoria_ram}")
    if equipo.almacenamiento:
        parts.append(f"Almacenamiento: {equipo.almacenamiento}")
    if equipo.mac_ethernet:
        parts.append(f"MAC Ethernet: {equipo.mac_ethernet}")
    if equipo.mac_wifi:
        parts.append(f"MAC WiFi: {equipo.mac_wifi}")
    return "<br/>".join(parts)


def build_safekeeping_pdf(equipo, fecha_asignacion):
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=f"Resguardo {equipo.numero_serie}",
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("NormalSICI", parent=styles["Normal"], fontName="Helvetica", fontSize=9, leading=12)
    centered = ParagraphStyle("CenteredSICI", parent=normal, alignment=TA_CENTER)
    justified = ParagraphStyle("JustifiedSICI", parent=normal, alignment=TA_JUSTIFY)
    small = ParagraphStyle("SmallSICI", parent=normal, fontSize=8, leading=10)
    title = ParagraphStyle("TitleSICI", parent=normal, fontName="Helvetica-Bold", fontSize=12, leading=15, alignment=TA_CENTER)

    location_label = "Mesa" if equipo.ubicacion.tipo == Ubicacion.Tipo.MESA else "Frente"
    location_name = equipo.ubicacion.nombre
    manager_role = equipo.ubicacion.responsable_cargo or ("Jefe de Mesa" if equipo.ubicacion.tipo == Ubicacion.Tipo.MESA else "I.R.O.")
    manager_name = equipo.ubicacion.responsable_nombre
    date_text = fecha_asignacion.strftime("%d/%m/%Y")

    story = [
        Paragraph("<b>Ejercito Mexicano.</b>", centered),
        Paragraph(f'Agto. Ings. "Felipe Angeles". &nbsp;&nbsp; {location_label} {location_name}.', centered),
        Spacer(1, 8),
        Paragraph("<b>RECIBO</b>", title),
        Spacer(1, 10),
        Paragraph('Unidad o dependencia: Agto. Ings. "Felipe Angeles".', normal),
        Paragraph(f"{location_label}: <b>{location_name}</b>.", normal),
        Paragraph(
            f"Ubicacion fisica: <b>{equipo.ubicacion_fisica.nombre if equipo.ubicacion_fisica else 'Santa Lucia'}</b>.",
            normal,
        ),
        Paragraph(f"Responsable: <b>{equipo.asignado_a}</b>.", normal),
        Paragraph(f"Concepto: Resguardo de equipos {location_label} {location_name}.", normal),
        Spacer(1, 10),
    ]

    equipment_table = Table(
        [
            [Paragraph("<b>Cantidad</b>", centered), Paragraph("<b>Descripcion</b>", centered), Paragraph("<b>Observaciones</b>", centered)],
            [Paragraph("1", centered), Paragraph(equipment_description_for_pdf(equipo), small), Paragraph(equipo.observaciones or "Sin observaciones.", small)],
        ],
        colWidths=[2 * cm, 11.7 * cm, 4.5 * cm],
        rowHeights=[0.8 * cm, 4.4 * cm],
    )
    equipment_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend(
        [
            equipment_table,
            Spacer(1, 12),
            Paragraph(
                f'Campo Mil. Estrategico Conjunto No. 37-D, "Gral. Div. P.A. Alfredo Lezama Alvarez", '
                f"Santa Lucia, Edo. de Mex., a {date_text}.",
                centered,
            ),
            Spacer(1, 12),
            Paragraph(
                "<b>Nota:</b> Los bienes antes descritos quedan bajo resguardo de la persona que los recibe, "
                "por lo que cualquier falla o descompostura ocasionada por el uso incorrecto de los mismos "
                "sera cubierta por dicha persona.",
                justified,
            ),
            Spacer(1, 28),
        ]
    )

    signature_table = Table(
        [
            [Paragraph("<b>Entrego:</b>", centered), Paragraph("<b>Recibio:</b>", centered)],
            [Spacer(1, 35), Spacer(1, 35)],
            [Paragraph("________________________________", centered), Paragraph("________________________________", centered)],
            [Paragraph("Nombre, grado y firma", centered), Paragraph(equipo.asignado_a, centered)],
        ],
        colWidths=[9.1 * cm, 9.1 * cm],
    )
    signature_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "BOTTOM")]))
    story.extend(
        [
            signature_table,
            Spacer(1, 26),
            Paragraph("<b>Autorizo:</b>", centered),
            Spacer(1, 34),
            Paragraph("________________________________", centered),
            Paragraph(f"{manager_role}.", centered),
            Paragraph(f"<b>{manager_name}</b>", centered),
        ]
    )

    document.build(story)
    buffer.seek(0)
    return buffer


# Generacion, sustitucion y consulta de evidencias firmadas.
@login_required
@require_POST
def generar_resguardo_equipo(request, numero_serie):
    equipo = get_object_or_404(
        Equipo.objects.select_related("ubicacion", "ubicacion_fisica").prefetch_related("movimientos"),
        numero_serie=numero_serie,
    )
    if not can_manage_equipment(request, equipo):
        return HttpResponse("No tienes permiso para generar este resguardo.", status=403)
    if equipo.estado != Equipo.Estado.ASIGNADO or not equipo.asignado_a:
        return HttpResponse("El equipo debe estar asignado para generar un resguardo.", status=400)

    responsable = current_responsable(request)
    pending_safekeeping = equipo.resguardos.filter(activo=True, archivo_firmado="").first()
    replace_pending = str(request.POST.get("reemplazar", "")).lower() in {"true", "1", "si"}
    if pending_safekeeping and not replace_pending:
        return JsonResponse(
            {
                "error": "Ya existe un resguardo generado pendiente de firma.",
                "requiresConfirmation": True,
            },
            status=409,
        )

    actor_name = responsable.nombre if responsable else request.user.get_full_name() or request.user.username
    if pending_safekeeping:
        pending_safekeeping.activo = False
        pending_safekeeping.reemplazado_por = request.user
        pending_safekeeping.reemplazado_por_nombre = actor_name
        pending_safekeeping.reemplazado_desde_ip = request_ip(request)
        pending_safekeeping.reemplazado = timezone.now()
        pending_safekeeping.save(
            update_fields=[
                "activo",
                "reemplazado_por",
                "reemplazado_por_nombre",
                "reemplazado_desde_ip",
                "reemplazado",
            ]
        )

    fecha_asignacion = assignment_date_for_equipment(equipo)
    resguardo = ResguardoEquipo.objects.create(
        equipo=equipo,
        asignado_a=equipo.asignado_a,
        ubicacion_nombre=equipo.ubicacion.nombre,
        ubicacion_fisica_nombre=equipo.ubicacion_fisica.nombre if equipo.ubicacion_fisica else "Santa Lucía",
        fecha_asignacion=fecha_asignacion,
        generado_por=request.user,
        generado_por_nombre=actor_name,
        generado_desde_ip=request_ip(request),
    )
    pdf = build_safekeeping_pdf(equipo, fecha_asignacion)
    response = FileResponse(pdf, as_attachment=True, filename=f"Resguardo-{equipo.numero_serie}.pdf")
    response["X-SICI-Resguardo-Id"] = str(resguardo.id)
    response["X-SICI-Resguardo-Fecha"] = resguardo.fecha_asignacion.strftime("%d/%m/%Y")
    response["X-SICI-Resguardo-Generado"] = timezone.localtime(resguardo.generado).strftime("%d/%m/%Y %H:%M")
    response["X-SICI-Resguardo-Usuario"] = quote(resguardo.generado_por_nombre)
    return response


@login_required
@require_POST
def cargar_resguardo_firmado(request, numero_serie):
    equipo = get_object_or_404(Equipo.objects.select_related("ubicacion"), numero_serie=numero_serie)
    if not can_manage_equipment(request, equipo):
        return JsonResponse({"error": "No tienes permiso para cargar este resguardo."}, status=403)

    uploaded_file = request.FILES.get("archivo")
    if not uploaded_file:
        return JsonResponse({"error": "Selecciona el PDF firmado."}, status=400)
    if uploaded_file.content_type != "application/pdf" or not uploaded_file.name.lower().endswith(".pdf"):
        return JsonResponse({"error": "El resguardo firmado debe ser un archivo PDF."}, status=400)
    if uploaded_file.size > 15 * 1024 * 1024:
        return JsonResponse({"error": "El PDF firmado no debe superar 15 MB."}, status=400)

    resguardo = equipo.resguardos.filter(activo=True, archivo_firmado="").first()
    if not resguardo:
        return JsonResponse({"error": "Primero genera el resguardo PDF desde la ficha."}, status=400)

    responsable = current_responsable(request)
    resguardo.archivo_firmado = uploaded_file
    resguardo.cargado_por = request.user
    resguardo.cargado_por_nombre = responsable.nombre if responsable else request.user.get_full_name() or request.user.username
    resguardo.cargado_desde_ip = request_ip(request)
    resguardo.cargado = timezone.now()
    resguardo.save(
        update_fields=[
            "archivo_firmado",
            "cargado_por",
            "cargado_por_nombre",
            "cargado_desde_ip",
            "cargado",
        ]
    )
    return JsonResponse({"safekeeping": safekeeping_payload(resguardo)})


@login_required
def descargar_resguardo_firmado(request, resguardo_id):
    resguardo = get_object_or_404(ResguardoEquipo.objects.select_related("equipo"), pk=resguardo_id)
    if not can_manage_equipment(request, resguardo.equipo):
        return HttpResponse("No tienes permiso para consultar este resguardo.", status=403)
    if not resguardo.archivo_firmado:
        return HttpResponse("Este resguardo no tiene un PDF firmado.", status=404)

    return FileResponse(
        resguardo.archivo_firmado.open("rb"),
        as_attachment=False,
        filename=f"Resguardo-firmado-{resguardo.equipo.numero_serie}.pdf",
        content_type="application/pdf",
    )


# Catalogo de tipos de equipo y atencion de solicitudes.
@login_required
@require_POST
def crear_tipo_equipo(request):
    if not request.user.is_staff:
        return JsonResponse({"error": "No tienes permiso para crear tipos de equipo."}, status=403)

    nombre = request.POST.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"error": "El nombre del tipo es obligatorio."}, status=400)

    tipo, created = TipoEquipo.objects.get_or_create(nombre__iexact=nombre, defaults={"nombre": nombre})
    if not created:
        return JsonResponse({"error": "Ya existe un tipo de equipo con ese nombre."}, status=400)

    return JsonResponse({"type": {"id": tipo.id, "nombre": tipo.nombre}}, status=201)


@login_required
@require_POST
def solicitar_tipo_equipo(request):
    nombre = request.POST.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"error": "El nombre del tipo es obligatorio."}, status=400)

    if TipoEquipo.objects.filter(nombre__iexact=nombre).exists():
        return JsonResponse({"error": "Ese tipo de equipo ya existe en el catalogo."}, status=400)

    if SolicitudTipoEquipo.objects.filter(
        nombre__iexact=nombre,
        estado=SolicitudTipoEquipo.Estado.PENDIENTE,
    ).exists():
        return JsonResponse({"error": "Ya existe una solicitud pendiente para ese tipo."}, status=400)

    solicitud = SolicitudTipoEquipo.objects.create(nombre=nombre, solicitante=request.user)
    return JsonResponse({"request": equipment_type_request_payload(solicitud)}, status=201)


@login_required
@require_POST
def aprobar_solicitud_tipo_equipo(request, solicitud_id):
    error = admin_required_json(request)
    if error:
        return error

    solicitud = get_object_or_404(SolicitudTipoEquipo, pk=solicitud_id)
    if solicitud.estado != SolicitudTipoEquipo.Estado.PENDIENTE:
        return JsonResponse({"error": "Esta solicitud ya fue atendida."}, status=400)

    tipo, _ = TipoEquipo.objects.get_or_create(nombre__iexact=solicitud.nombre, defaults={"nombre": solicitud.nombre})
    solicitud.estado = SolicitudTipoEquipo.Estado.APROBADA
    solicitud.save()
    return JsonResponse({
        "request": equipment_type_request_payload(solicitud),
        "type": {"id": tipo.id, "nombre": tipo.nombre},
    })


@login_required
@require_POST
def rechazar_solicitud_tipo_equipo(request, solicitud_id):
    error = admin_required_json(request)
    if error:
        return error

    solicitud = get_object_or_404(SolicitudTipoEquipo, pk=solicitud_id)
    if solicitud.estado != SolicitudTipoEquipo.Estado.PENDIENTE:
        return JsonResponse({"error": "Esta solicitud ya fue atendida."}, status=400)

    solicitud.estado = SolicitudTipoEquipo.Estado.RECHAZADA
    solicitud.comentario_admin = request.POST.get("comentario", "").strip()
    solicitud.save()
    return JsonResponse({"request": equipment_type_request_payload(solicitud)})


# Catalogo de marcas y atencion de solicitudes.
@login_required
@require_POST
def crear_marca_equipo(request):
    error = admin_required_json(request)
    if error:
        return error

    nombre = request.POST.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"error": "El nombre de la marca es obligatorio."}, status=400)

    marca, created = MarcaEquipo.objects.get_or_create(nombre__iexact=nombre, defaults={"nombre": nombre})
    if not created:
        return JsonResponse({"error": "Ya existe una marca con ese nombre."}, status=400)

    return JsonResponse({"brand": {"id": marca.id, "nombre": marca.nombre}}, status=201)


@login_required
@require_POST
def solicitar_marca_equipo(request):
    nombre = request.POST.get("nombre", "").strip()
    if not nombre:
        return JsonResponse({"error": "El nombre de la marca es obligatorio."}, status=400)

    if MarcaEquipo.objects.filter(nombre__iexact=nombre).exists():
        return JsonResponse({"error": "Esa marca ya existe en el catalogo."}, status=400)

    if SolicitudMarcaEquipo.objects.filter(
        nombre__iexact=nombre,
        estado=SolicitudMarcaEquipo.Estado.PENDIENTE,
    ).exists():
        return JsonResponse({"error": "Ya existe una solicitud pendiente para esa marca."}, status=400)

    solicitud = SolicitudMarcaEquipo.objects.create(nombre=nombre, solicitante=request.user)
    return JsonResponse({"request": equipment_brand_request_payload(solicitud)}, status=201)


@login_required
@require_POST
def aprobar_solicitud_marca_equipo(request, solicitud_id):
    error = admin_required_json(request)
    if error:
        return error

    solicitud = get_object_or_404(SolicitudMarcaEquipo, pk=solicitud_id)
    if solicitud.estado != SolicitudMarcaEquipo.Estado.PENDIENTE:
        return JsonResponse({"error": "Esta solicitud ya fue atendida."}, status=400)

    marca, _ = MarcaEquipo.objects.get_or_create(nombre__iexact=solicitud.nombre, defaults={"nombre": solicitud.nombre})
    solicitud.estado = SolicitudMarcaEquipo.Estado.APROBADA
    solicitud.save()
    return JsonResponse({
        "request": equipment_brand_request_payload(solicitud),
        "brand": {"id": marca.id, "nombre": marca.nombre},
    })


@login_required
@require_POST
def rechazar_solicitud_marca_equipo(request, solicitud_id):
    error = admin_required_json(request)
    if error:
        return error

    solicitud = get_object_or_404(SolicitudMarcaEquipo, pk=solicitud_id)
    if solicitud.estado != SolicitudMarcaEquipo.Estado.PENDIENTE:
        return JsonResponse({"error": "Esta solicitud ya fue atendida."}, status=400)

    solicitud.estado = SolicitudMarcaEquipo.Estado.RECHAZADA
    solicitud.comentario_admin = request.POST.get("comentario", "").strip()
    solicitud.save()
    return JsonResponse({"request": equipment_brand_request_payload(solicitud)})
